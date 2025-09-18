import openpyxl, re, os
import mysql.connector
from datetime import date

EXCEL_PATH = "/opt/data/casting.xlsx"
SHEET_NAME = "DailyPerformance"  # exactly as listed

# headers we expect (case-insensitive contains)
HDR_LABELS = ["row", "labels"]
HDR_GOOD   = ["sum of total good parts", "good parts", "total good"]
HDR_OEE    = ["sum of ind. oee", "oee"]

def find_header_cols(ws):
    row1 = {c.column_letter: (str(c.value).strip().lower() if c.value else "") for c in ws[1]]
    # fallback: scan first 5 rows for a header row
    for r in range(1, 6):
        row = {c.column_letter: (str(c.value).strip().lower() if c.value else "") for c in ws[r]]
        txt = " ".join(row.values())
        if "labels" in txt and "oee" in txt:
            row1 = row; break

    def pick(colmap, keywords):
        for col, txt in colmap.items():
            s = txt.replace("(pcs)", "")
            if any(k in s for k in keywords):
                return col
        return None

    label_col = pick(row1, HDR_LABELS)
    good_col  = pick(row1, HDR_GOOD)
    oee_col   = pick(row1, HDR_OEE)
    return label_col, good_col, oee_col

def parse_number(v):
    if v is None: return None
    if isinstance(v,(int,float)): return float(v)
    s = str(v).replace(",", "").strip()
    if not s or s.upper().startswith("#DIV"): return None
    try: return float(s)
    except: return None

def get_group_or_machine(txt):
    s = (txt or "").strip()
    if not s: return None, None
    if s.startswith("#"):               # group line like "#05 - 800T to 850T"
        return s, None
    m = re.search(r"\bDCM#\d+\b", s, re.I)
    if m:
        return None, m.group(0).upper() # machine code
    return None, None

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Sheet '{SHEET_NAME}' not found. Found: {wb.sheetnames}")
        return
    ws = wb[SHEET_NAME]

    label_col, good_col, oee_col = find_header_cols(ws)
    if not all([label_col, good_col, oee_col]):
        print("Could not detect header columns. Check the sheet headers.")
        return
    print(f"Using columns: label={label_col}, good={good_col}, oee={oee_col}")

    rows=[]
    current_group=None
    for r in range(2, ws.max_row+1):
        label = ws[f"{label_col}{r}"].value
        good  = parse_number(ws[f"{good_col}{r}"].value)
        oee   = parse_number(ws[f"{oee_col}{r}"].value)

        if not label:
            continue
        # stop at Grand Total
        if isinstance(label, str) and label.strip().lower().startswith("grand total"):
            break

        grp, machine = get_group_or_machine(str(label))
        if grp:
            current_group = grp
            continue
        if not machine:
            # skip part rows like '2238#1'
            continue

        rows.append((current_group or "", machine, int(good or 0), (oee if oee is not None else None)))

    if not rows:
        print("No machine rows parsed.")
        return

    snap = date.today()
    print(f"Parsed {len(rows)} machines; snapshot_date={snap}")

    conn = mysql.connector.connect(host="db", user="root", password="secret", database="casting")
    cur = conn.cursor()
    cur.executemany("""
        REPLACE INTO machine_perf_snapshot (snapshot_date, group_label, machine_code, total_good_parts, oee_pct)
        VALUES (%s,%s,%s,%s,%s)
    """, [(snap, g, m, gp, o) for (g,m,gp,o) in rows])
    conn.commit(); cur.close(); conn.close()
    print("Inserted/updated snapshot successfully.")

if __name__ == "__main__":
    main()
