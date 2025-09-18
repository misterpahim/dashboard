import os, re
import openpyxl
import mysql.connector
from datetime import datetime, date
from openpyxl.utils.datetime import from_excel as xl_from_serial

EXCEL_PATH = "/opt/data/casting.xlsx"

DAILY_CELLS = [
    ("OEE_PCT",        "Z172",  True),
    ("TOTAL_UP_MIN",   "BD172", False),
    ("DIE_DOWN_MIN",   "BF172", False),
    ("MACH_DOWN_MIN",  "BH172", False),
    ("ENG_DOWN_MIN",   "BJ172", False),
    ("PROD_DOWN_MIN",  "BL172", False),
    ("OTH_DOWN_MIN",   "BN172", False),
    ("TOTAL_DOWN_MIN", "BP172", False),
]
ERR = {"#VALUE!","#DIV/0!","#N/A","#NAME?","#REF!","#NULL!","#NUM!"}

def safe_number(v, pct=False):
    if v is None: return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s.upper() in ERR: return None
        try: v = float(s)
        except: return None
    try:
        v = float(v)
        return v*100.0 if pct else v
    except: return None

def coerce_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, (int, float)):
        try: return xl_from_serial(v).date()
        except: return None
    if isinstance(v, str):
        for fmt in ("%d-%b-%y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try: return datetime.strptime(v, fmt).date()
            except: pass
    return None

def parse_month_year_from_filename(path):
    name = os.path.basename(path)
    # accept: '08 -2025', '08-2025', '2025-08', 'Aug 2025'
    m = re.search(r'(\b\d{2})\s*-\s*(\d{4}\b)', name)           # MM-YYYY
    if m: return int(m.group(2)), int(m.group(1))
    m = re.search(r'(\b\d{4})\s*-\s*(\d{2}\b)', name)           # YYYY-MM
    if m: return int(m.group(1)), int(m.group(2))
    m = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+(\d{4})', name, re.I)
    if m:
        mm = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}[m.group(1)[:3].lower()]
        return int(m.group(2)), mm
    return None

def main():
    print("Loading workbook (read_only, data_only)…")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)

    ym = parse_month_year_from_filename(EXCEL_PATH)  # fallback month/year
    if not ym:
        print("ℹ️  Filename month/year not found; will rely on D6 per sheet and skip if missing.")

    rows = []
    for sn in wb.sheetnames:
        sn_clean = sn.strip()
        if not sn_clean.isdigit():
            continue
        ws = wb[sn]

        # 1) prefer real date from D6
        d = coerce_date(ws["D6"].value)
        # 2) fallback to filename month/year + sheet day
        if not d and ym:
            yy, mm = ym
            try:
                d = date(yy, mm, int(sn_clean))
            except ValueError:
                d = None

        if not d:
            print(f"[{sn}] no usable date (D6 empty and filename fallback unavailable) — skipping.")
            continue

        print(f"Reading sheet {sn} -> report_date {d}")
        for code, cell, is_pct in DAILY_CELLS:
            val = safe_number(ws[cell].value, pct=is_pct)
            print(f"  {code} ({cell}): {val}")
            if val is not None:
                rows.append((d, code, val))

    print(f"Collected {len(rows)//len(DAILY_CELLS) if rows else 0} days of data.")
    if not rows:
        print("No data found."); return

    conn = mysql.connector.connect(host="db", user="root", password="secret", database="casting")
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS kpi_facts(
        report_date DATE NOT NULL,
        metric_code VARCHAR(64) NOT NULL,
        metric_value DOUBLE,
        PRIMARY KEY(report_date, metric_code)
    )""")
    cur.execute("""CREATE OR REPLACE VIEW v_daily_oee AS
      SELECT
        MAX(CASE WHEN metric_code='TOTAL_UP_MIN'   THEN metric_value END) AS total_up_min,
        MAX(CASE WHEN metric_code='DIE_DOWN_MIN'   THEN metric_value END) AS die_down_min,
        MAX(CASE WHEN metric_code='MACH_DOWN_MIN'  THEN metric_value END) AS mach_down_min,
        MAX(CASE WHEN metric_code='ENG_DOWN_MIN'   THEN metric_value END) AS eng_down_min,
        MAX(CASE WHEN metric_code='PROD_DOWN_MIN'  THEN metric_value END) AS prod_down_min,
        MAX(CASE WHEN metric_code='OTH_DOWN_MIN'   THEN metric_value END) AS oth_down_min,
        MAX(CASE WHEN metric_code='TOTAL_DOWN_MIN' THEN metric_value END) AS total_down_min,
        MAX(CASE WHEN metric_code='OEE_PCT'        THEN metric_value END) AS oee_pct,
        report_date
      FROM kpi_facts
      GROUP BY report_date""")
    cur.executemany("REPLACE INTO kpi_facts(report_date,metric_code,metric_value) VALUES(%s,%s,%s)", rows)
    conn.commit(); cur.close(); conn.close()
    print(f"Inserted {len(rows)} rows successfully.")

if __name__=="__main__":
    main()
